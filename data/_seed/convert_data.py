"""Convert the Excel dataset to JSON for the web application."""
import json
import re
import openpyxl

wb = openpyxl.load_workbook("Database-Bugatti-Miniatures.xlsx")
ws = wb["Full"]

HEADERS = [
    "fabricant", "ref", "serie", "marque", "type_bugatti",
    "modele", "chassis", "annee", "couleur", "echelle",
    "type_miniature", "annee_miniature", "remarques",
    "source_photo", "source_info", "montage"
]

miniatures = []
for row_idx in range(2, ws.max_row + 1):
    raw = [cell.value for cell in ws[row_idx]]
    entry = {}
    for i, h in enumerate(HEADERS):
        val = raw[i]
        if val is not None:
            entry[h] = str(val).strip() if not isinstance(val, (int, float)) else val
    if not entry.get("fabricant") and not entry.get("type_bugatti"):
        continue
    entry["id"] = row_idx - 1
    # Normalize echelle
    if "echelle" in entry:
        entry["echelle"] = f"1/{entry['echelle']}"
    # Extract type number for grouping
    tb = entry.get("type_bugatti", "")
    m = re.match(r"^(\d+)", str(tb))
    entry["type_number"] = m.group(1) if m else ""
    miniatures.append(entry)

# Build stats
fabricants_count = {}
types_count = {}
materials_count = {}
decades = {}
marques_count = {}

for m in miniatures:
    fab = m.get("fabricant", "Inconnu")
    fabricants_count[fab] = fabricants_count.get(fab, 0) + 1

    tn = m.get("type_number", "Autre")
    if tn:
        types_count[tn] = types_count.get(tn, 0) + 1

    mat = m.get("type_miniature", "Non spécifié")
    materials_count[mat] = materials_count.get(mat, 0) + 1

    annee = str(m.get("annee_miniature", ""))
    if annee.isdigit():
        decade = f"{annee[:3]}0s"
        decades[decade] = decades.get(decade, 0) + 1

    marque = m.get("marque", "Bugatti")
    marques_count[marque] = marques_count.get(marque, 0) + 1

# Top fabricants
top_fabricants = sorted(fabricants_count.items(), key=lambda x: -x[1])[:50]
top_types = sorted(types_count.items(), key=lambda x: -x[1])[:40]

data = {
    "miniatures": miniatures,
    "stats": {
        "total": len(miniatures),
        "fabricants": len(fabricants_count),
        "types": len(types_count),
        "top_fabricants": top_fabricants,
        "top_types": top_types,
        "materials": sorted(materials_count.items(), key=lambda x: -x[1]),
        "decades": sorted(decades.items()),
        "marques": sorted(marques_count.items(), key=lambda x: -x[1]),
    },
}

with open("static/data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=None)

print(f"Exported {len(miniatures)} miniatures")
print(f"Stats: {len(fabricants_count)} fabricants, {len(types_count)} types")
